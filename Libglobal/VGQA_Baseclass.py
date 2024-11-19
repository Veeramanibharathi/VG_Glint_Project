from selenium import webdriver
from selenium.webdriver import ChromeOptions, ActionChains
import openpyxl
import datetime

from selenium.webdriver.support.select import Select


class Utility:
    def open_browser(self):
        chrome_option = ChromeOptions()
        chrome_option.add_experimental_option("detach", True)
        self.driver = webdriver.Chrome(options=chrome_option)
        return self.driver

    # WebDriver:C
    def window_maximize(self):
        self.driver.maximize_window()

    def load_url(self, url):
        self.driver.get(url)

    def btn_click(self, ele):
        ele.click()

    def close_browser(self):
        self.driver.quit()

    def java_script_executor(self, script, *element):
        self.driver.execute_script(script, element)

    def page_screenshot(self, filename):
        self.driver.save_screenshot(filename)

    def close_tab(self):
        self.driver.close()

    def forward_page(self):
        self.driver.forward()

    def back_page(self):
        self.driver.back()

    def refresh_page(self):
        self.driver.refresh()

    def implicit_wait(self, sec):
        self.driver.implicitly_wait(sec)

    def get_page_url(self):
        url = self.driver.current_url
        return url

    def get_page_title(self):
        title = self.driver.title
        return title

    def get_parent_window_id(self):
        par_window_id = self.driver.current_window_handle
        return par_window_id

    def get_all_windows_id(self):
        all_windows_id = self.driver.window_handles
        return all_windows_id

    def switch_to_element(self):
        switch_to = self.driver.switch_to
        return switch_to

    # Excel Read
    def get_data_from_excel(self, sheet_name, row_no, cell_no):
        loc = r"C:\Users\nitis\PycharmProjects\TekVision\TestData\Data.xlsx"
        w = openpyxl.load_workbook(loc)
        sheet = w.get_sheet_by_name(sheet_name)
        cell = sheet.cell(row_no, cell_no)
        value = cell.value
        if type(value) == datetime.datetime:
            date = value.strftime("%d/%m/%Y")
            return date
        else:
            return value

    # Excel Update
    def update_date_in_excel(self, sheet, row_now, cell_now, value):
        loc = ""
        w = openpyxl.Workbook()
        sheet = w.create_sheet(sheet)
        cell = sheet.cell(row_now, cell_now)
        cell.value = value
        w.save(loc)

    # WebElement:C
    def insert_value(self, element, data):
        element.send_keys(data)

    def element_click(self, element):
        element.click()

    def clear_value(self, element):
        element.clear()

    def get_attribute_value(self, element, name):
        attribute = element.get_attribute(name)
        return attribute

    def displayed(self, element):
        d = element.is_displayed()
        return d

    def enabled(self, element):
        e = element.is_enabled()
        return e

    def selected(self, element):
        s = element.is_selected()
        return s

    def get_element_text(self, element):
        text = element.text
        return text

    # Select:C
    def select_option_by_index(self, element, index):
        s = Select(element)
        s.select_by_index(index)

    def select_option_by_value(self, element, value):
        s = Select(element)
        s.select_by_value(value)

    def select_option_by_text(self, element, text):
        s = Select(element)
        s.select_by_visible_text(text)

    def deselect_option_by_index(self, element, index):
        s = Select(element)
        s.deselect_by_index(index)

    def deselect_option_by_value(self, element, value):
        s = Select(element)
        s.deselect_by_value(value)

    def deselect_option_by_text(self, element, text):
        s = Select(element)
        s.deselect_by_visible_text(text)

    def deselect_all_values(self, element):
        s = Select(webelement=element)
        s.deselect_all()

    def get_options(self, element):

        s = Select(webelement=element)
        options = s.options
        return options

    def select_all_option_by_index(self, element):
        s = Select(element)
        options = self.get_options(element=element)
        for index, each_option in enumerate(options):
            s.select_by_index(index)

    def select_all_option_by_text(self, element):
        s = Select(element)
        options = self.get_options(element=element)
        for each_option in options:
            element_text = self.get_element_text(element=each_option)
            self.select_option_by_text(element=element, text=element_text)

    def select_all_option_by_value(self, element):
        s = Select(element)
        options = self.get_options(element=element)
        for each_option in options:
            value = self.get_attribute_value(element=each_option, name="value")
            self.select_option_by_value(element=element, value=value)

    def mouse_over(self, element):
        a = ActionChains(self.driver)
        a.move_to_element(element).perform()

    def right_click(self, element):
        a = ActionChains(self.driver)
        a.context_click(element).perform()

    def drag_and_drop(self, source, target):
        a = ActionChains(self.driver)
        a.drag_and_drop(source, target).perform()

    def double_click(self, element):
        a = ActionChains(self.driver)
        a.double_click(element).perform()

    def switch_to_alert(self):
        al = self.switch_to_element().alert
        return al

    def accept_alert(self):
        self.switch_to_alert().accept()

    def dismiss_alert(self):
        self.switch_to_alert().dismiss()

    def enter_text_in_alert(self, data):
        self.switch_to_alert().send_keys(data)

    def get_alert_text(self):
        text = self.switch_to_alert().text
        return text

    def switch_to_window_by_id(self, id):
        self.switch_to_element().window(id)

    def switch_to_window_by_url(self, url):
        self.switch_to_element().window(url)

    def switch_to_window_by_title(self, title):
        self.switch_to_element().window(title)

    def switch_to_particular_window_by_index(self, index):
        all_windows_id = self.get_all_windows_id()
        self.switch_to_element().window(all_windows_id[index])

    def switch_to_parent_window(self):
        parent_window_id = self.get_parent_window_id()
        self.switch_to_element().window(parent_window_id)

    def switch_to_frame_by_id(self, id):
        self.switch_to_element().frame(id)

    def switch_to_frame_by_name(self, name):
        self.switch_to_element().frame(name)

    def switch_to_frame_by_element(self, element):
        self.switch_to_element().frame(element)

    def switch_to_frame_by_index(self, index):
        self.switch_to_element().frame(index)

    def switch_to_parent_frame(self):
        self.switch_to_element().parent_frame()

    def insert_value_using_java_script(self, data, *element):
        self.java_script_executor("arguments[0].setAttribute('value','{}')".format(data), element)

    def click_using_java_script(self, *element):
        self.java_script_executor("arguments[0].click()", element)

    def get_attribute_value_using_java_script(self, *element):
        self.java_script_executor("arguments[0].getAttribute('value')", element)

    def scroll_down(self, *element):
        self.java_script_executor("arguments[0].scrollIntoView(true)", element)

    def scroll_up(self, *element):
        self.java_script_executor("arguments[0].scrollIntoView(false)", element)

